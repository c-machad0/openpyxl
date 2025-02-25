from openpyxl import load_workbook
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList

file_path = 'estoque.xlsx'
workbook = load_workbook(file_path)

sheet = workbook['Estoque']

graph_sheet = workbook.create_sheet(title='Gráficos')

max_row = sheet.max_row

col_nome_produto = 1 # Coluna A
col_valor_fornecedor = 2 # Coluna B
col_lucratividade = 3 # Coluna C
col_quantidade = 4 # Coluna D
col_preco_venda = 5 # Coluna E
col_lucro_total = 6 # Coluna F
col_valor_total = 7 # Coluna G

if sheet.cell(row=max_row, column=col_nome_produto).value == 'Totais Gerais':
    max_row -= 2

bar_chart_valor_total = BarChart()
bar_chart_valor_total.title = 'Valor Total em Estoque por Produto'
bar_chart_valor_total.y_axis.title = 'Valor Total (R$)'
bar_chart_valor_total.x_axis.title = 'Produto'

data = Reference(sheet, min_col=col_valor_total, min_row=1, max_row=max_row)
cats = Reference(sheet, min_col=col_valor_total, min_row=2, max_row=max_row)

bar_chart_valor_total.add_data(data, titles_from_data=True)
bar_chart_valor_total.set_categories(cats)

bar_chart_valor_total.width = 30
bar_chart_valor_total.height = 15

graph_sheet.add_chart(bar_chart_valor_total, 'A1')

produtos_lucratividade = []

for row in range(2, max_row + 1):
    nome_produto = sheet.cell(row=row, column=col_nome_produto).value
    lucratividade_cell = sheet.cell(row=row, column=col_lucratividade)
    lucratividade = lucratividade_cell.value

    if lucratividade is not None and isinstance(lucratividade, (int, float)):
        produtos_lucratividade.append((nome_produto, lucratividade))

if produtos_lucratividade:
    produtos_lucratividade.sort(key=lambda x: x[1], reverse=True)
    
    top_5_lucrativos = produtos_lucratividade[:5]

    aux_sheet = workbook.create_sheet(title='Auxiliar')
    aux_sheet.cell(row=1, column=1, value='Nome do Produto')
    aux_sheet.cell(row=1, column=2, value='Lucratividade (%)')

    for idx, (nome_produto, lucratividade) in enumerate(top_5_lucrativos, start=2):
        aux_sheet.cell(row=idx, column=1, value=nome_produto)
        aux_sheet.cell(row=idx, column=2, value=lucratividade)

    pie_chart_lucratividade = PieChart()
    pie_chart_lucratividade.title = 'Tops 5 Produtos com Maior Lucrativdade'

    data = Reference(aux_sheet, min_col=2, min_row=2, max_row=6)
    labels = Reference(aux_sheet, min_col=1, min_row=2, max_row=6)

    pie_chart_lucratividade.add_data(data,titles_from_data=False)
    pie_chart_lucratividade.set_categories(labels)

    pie_chart_lucratividade.dataLabels = DataLabelList()
    pie_chart_lucratividade.dataLabels.showVal = True
    pie_chart_lucratividade.dataLabels.showPercent = True

    pie_chart_lucratividade.width = 20
    pie_chart_lucratividade.height = 15

    graph_sheet.add_chart(pie_chart_lucratividade, 'A25')

line_chart_quantidade = LineChart()
line_chart_quantidade.title = 'Quantidade de Produtos em Estoque'
line_chart_quantidade.y_axis.title = 'Quantidade'
line_chart_quantidade.x_axis.title = 'Produto'

data = Reference(sheet, min_col=col_quantidade, min_row=1, max_row=max_row)
cats = Reference(sheet, min_col=col_nome_produto, min_row=2, max_row=max_row)

line_chart_quantidade.add_data(data, titles_from_data=True)
line_chart_quantidade.set_categories(cats)

line_chart_quantidade.width = 30
line_chart_quantidade.height = 15

graph_sheet.add_chart(line_chart_quantidade, 'A50')

workbook.save(file_path)