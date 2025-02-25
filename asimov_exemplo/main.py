from openpyxl import Workbook
import random

workbook = Workbook() # Instanciar uma classe

sheet = workbook.active # Ativando a nossa classe e criando na memória
sheet.title = 'Estoque'

headers = ['Nome do produto', 'Valor do fornecedor', 'Lucratividade (%)', 'Quantidade']

for col_num, header in enumerate(headers, start=1):
    sheet.cell(row=1, column=col_num, value=header) # Colocando os cabeçalhos logo na primeira linha

def gerar_nome_produto():
    prefixos = ['Super', 'Mega', 'Ultra', 'Power', 'Eco', 'Max']
    tipos = ['Widget', 'Gadget', 'Device', 'Tool', 'Instrument', 'Appliance']
    sufixos = ['Plus', 'Pro', 'X', '2000', 'Prime', 'Elite']
    return f'{random.choice(prefixos)} {random.choice(tipos)} {random.choice(sufixos)}'

num_produtos = 50

for row_num in range(2, num_produtos + 2):
    nome_produto = gerar_nome_produto()
    valor_fornecedor = round(random.uniform(10.0, 500.0), 2)
    lucratividade = random.randint(10, 100)
    quantidade = random.randint(1, 100)

    sheet.cell(row=row_num, column=1, value=nome_produto)
    sheet.cell(row=row_num, column=2, value=valor_fornecedor)
    sheet.cell(row=row_num, column=3, value=lucratividade)
    sheet.cell(row=row_num, column=4, value=quantidade)

file_path = 'estoque.xlsx'
workbook.save(file_path)
    

