from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

wb = load_workbook('dados.xlsx', data_only=True)
sheet = wb['Vendas']

valores_de_vendas = [] # Inicializa a lista que armazenará os valores das vendas de cada produto

for row in sheet.iter_rows(min_col=4, max_row=sheet.max_row, min_row=2, max_col=4): # Itera sobre as linhas da ultima coluna
    for cell in row: # Itera sobre as células de cada linha da coluna correspondente
        if isinstance(cell.value, (int, float)):
            valores_de_vendas.append(cell.value) # Adiciona o valor da célula na lista

total_de_vendas = sum(valores_de_vendas) # Soma todos os valores

max_produto = 0 # Inicializa a variavel que armazenará a quantidade de produtos
produto_mais_vendido = "" # Inicializa a variavel que armazenará o nome do produto mais vendido

# Itera sobre a linha dentro do intervalo estabelecido
# Como a linha gera uma tupla, pegar o primeiro valor da tupla [0] é o nome do produto e o segundo [1] é a quantidade
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=2):
    nome_produto = row[0].value # Nome do produto
    vendas = row[1].value # Quantidade vendida

    if vendas and isinstance(vendas, (int, float)): # Verifica se 'vendas' não é None e se o conteudo é int ou float
        if vendas > max_produto: # Se vendas for maior que a quantidade armazenada em max_produto
            max_produto = vendas # max_produto agora armazena a nova quantidade de vendas maior
            produto_mais_vendido = nome_produto # E o produto mais vendido se torna o produto com a maior quantidade de vendas

produto_mais_rentavel = "" # Armazena o nome do produto mais rentável
total_produto_mais_rentavel = 0 # Armazena o valor de vendas do produto mais rentável

for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=4): # Itera sobre as linhas dentro do intervalo da primeira coluna até a ultima coluna
    nome_produto_mais_rentavel = row[0].value # Pega o valor da celula que contem o nome do produto
    receita_total = row[3].value # Pega o valor da celula que contem a rentabilidade do produto

    if receita_total and isinstance(receita_total, (int, float)):
        if receita_total > total_produto_mais_rentavel:
            total_produto_mais_rentavel = receita_total
            produto_mais_rentavel = nome_produto_mais_rentavel

valores = [total_de_vendas, produto_mais_vendido, produto_mais_rentavel] # Armazena uma lista que servirá como limite do for

if 'Relatorio' in wb.sheetnames: # Se ja existir uma aba com o nome relatorio, ativa essa aba e escreve os valores. Se não, cria a aba
    sheet_destino = wb['Relatorio']

    for row in range(2, len(valores) + 1):
        sheet_destino.cell(row=row, column=1, value=total_de_vendas)
        sheet_destino.cell(row=row, column=2, value=produto_mais_vendido)
        sheet_destino.cell(row=row, column=3, value=produto_mais_rentavel)

else:
    sheet_destino = wb.create_sheet('Relatorio')

    headers = ['Total de vendas', 'Produto mais vendido', 'Produto mais rentável']

    for id, header in enumerate(headers, start=1):
        sheet_destino.cell(row=1, column=id, value=header)

    for col in sheet_destino.iter_cols(max_row=1, max_col=sheet_destino.max_column, min_col=1):
        coll_letter = col[0].column_letter
        sheet_destino.column_dimensions[coll_letter].width = 23

        for cell in col:
            cell.font = Font(size=12, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color='FF0000', fill_type='solid')

    for row in range(2, len(valores)):
        cell_total_vendas = sheet_destino.cell(row=row, column=1, value=total_de_vendas)
        cell_produto_mais_vendido = sheet_destino.cell(row=row, column=2, value=produto_mais_vendido)
        cell_total_vendas = sheet_destino.cell(row=row, column=3, value=produto_mais_rentavel)

wb.save('dados.xlsx')