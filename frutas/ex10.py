from openpyxl import load_workbook
import random

def cria_nome_produto(produtos_disponiveis):
    return random.choice(produtos_disponiveis)

wb = load_workbook('dados.xlsx')

sheets = ['Janeiro', 'Fevereiro', 'Março', 'Resumo']
headers = ['Produto', 'Quantidade', 'Preço Unitário', 'Total']

for sheet in sheets: # Itera sobre as abas
    ws = wb.create_sheet(f'{sheet}') # Cria cada aba inicia o preenchimento no proximo laço
    for idx, header in enumerate(headers, start=1): # Itera sobre a aba atual e cria os cabeçalhos
        ws.cell(row=1, column=idx, value=header) # Escreve os cabeçalhos

    if sheet != 'Resumo': # Verifica se a aba do laço é 'Resumo', pois não iremos preencher as linhas e colunas dela
        produtos = ['Uva', 'Morango', 'Maçã', 'Manga', 'Kiwi']

        for row in range(2, 2 + len(produtos)): # Itera sobre cada linha, iniciando pela linha 2 e terminando em len(produtos), ou seja, da linha 2 à linha 6
            # como len(produtos) = 5, soma-se 2 a ela pois estamos iniciando na linha 2 e porque o range não captura o ultimo valor, então para que a linha 6
            # seja capturada, precisamos iterar da linha 2, 2 + 5 = 7
            nome_produto = cria_nome_produto(produtos)
            produtos.remove(nome_produto)
            quantidade_produto = random.randint(1, 20)
            preco_produto = round(random.uniform(1, 10), 2)

            ws.cell(row=row, column=1, value=nome_produto)
            ws.cell(row=row, column=2, value=quantidade_produto)
            ws.cell(row=row, column=3, value=preco_produto)
 
wb.save('dados.xlsx')