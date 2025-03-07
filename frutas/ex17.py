from openpyxl import load_workbook

wb = load_workbook('dados.xlsx')

ws = wb['Resumo']

ws.cell(row=1, column=3, value='Preço Médio Unitário') # Atualizando o titulo da coluna 3

for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=ws.max_column): # Iterando sobre todas as colunas
    col_letter = col[0].column_letter # Pegando a letra da coluna atual do laço
    max_length = 0 # Inicializando variavel que armazenará o tamanho
    
    for cell in col: # Itera sobre cada célula da coluna
        if cell.value: # Se não for 'None'
            max_length = max(max_length, len(str(cell.value))) # Substitui o valor de max_length pelo tamanho da string da célula atual, até pegar o tamanho da maior string

    ws.column_dimensions[col_letter].width = max_length + 4 # Utiliza o tamanho da maior string como parâmetro para formatar a largura da celula com + 4 de espaçamento

wb.save('dados.xlsx')