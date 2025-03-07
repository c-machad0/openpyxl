from openpyxl import load_workbook

wb = load_workbook('dados.xlsx')

ws = wb['Resumo']

for row in range(2, 1 + ws.max_row): # Itera sobre todas as linhas e colunas da aba Resumo
    cell_quantidade = ws.cell(row=row, column=2) # Armazena as celulas da linha de quantidade
    cell_preco_unitario = ws.cell(row=row, column=3) # Armazena as celulas da linha de preço unitário
    cell_total_vendas = ws.cell(row=row, column=4) # Armazena as celulas da linha de total

    formula_total_vendas = f'={cell_quantidade.coordinate}*{cell_preco_unitario.coordinate}' # Armazena a formula que multiplica a quantidade * preço unitario de cada linha -> B1 * C1
    cell_total_vendas.value = formula_total_vendas # Pega o valor da formula

wb.save('dados.xlsx')
