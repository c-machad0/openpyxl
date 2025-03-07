from openpyxl import load_workbook

wb = load_workbook('dados.xlsx')

sheets = ['Janeiro', 'Fevereiro', 'Março']
sheet_resumo = wb['Resumo']

preco_por_produto = {}

for sheet in sheets:
    ws = wb[sheet]

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3): # Itera sobre as linhas, buscando o intervalo da coluna 'Produto' e 'Preço unitário'
        produto = row[0].value # Nome do produto
        preco_unitario = row[2].value # Preço unitário

        if isinstance(preco_unitario, (int, float)): # Verifica se é numero
            if produto in preco_por_produto: # Verifica se o produto está no dicionário
                preco_por_produto[produto] = preco_por_produto[produto] + preco_unitario # Se estiver, pega o valor do preço unitario daquele produto e soma com o preço atual
            else:
                preco_por_produto[produto] = preco_unitario # Se não estiver, adiciona o produto com o valor do laço atual

produtos_ordenados = sorted(preco_por_produto.keys()) # Ordena os produtos por ordem alfabética
linha_resumo = 2 # Inicializa a linha em 2, pois a 1 é a de cabeçalho

for nome_produto in produtos_ordenados: # Itera sob os produtos ordenados
    sheet_resumo.cell(row=linha_resumo, column=3, value=preco_por_produto[nome_produto]/3) # Escreve linha por linha, na coluna de preço unitário, o preço unitário médio do produto
    # Por isso preco_por_produto[nome_produto] / 3, pois o '3' refere-se a quantidade de meses calculado
    linha_resumo += 1 # Segue para a proxima linha assim que a instrução acima, acaba

wb.save('dados.xlsx')

"""
Para saber o preço médio de outra forma, se caso fosse usado uma outra lógica
for nome_produto, valor_por_produto in preco_por_produto.items():
    novo_valor = preco_por_produto[nome_produto] / 3
    preco_por_produto[nome_produto] = novo_valor
"""