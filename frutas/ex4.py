from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

wb = load_workbook('dados.xlsx')
ws = wb.active

for col in ws.iter_cols(min_col=1, max_row=1, max_col=ws.max_column): # Itera sobre todas as colunas
    for cell in col: # Itera sobre cada celula da linha do cabeçalho
        cell.font = Font(size=12, bold=True) # Altera a fonte para tamanho 12 e negrito
        cell.fill = PatternFill(start_color='FF0000', fill_type='solid') # Altera o preenchimento para cor vermelha
        cell.alignment = Alignment(horizontal='center', vertical='center') # Altera o alinhamento para o centro da célula

for col in ws.iter_cols(min_col=1, max_row=1, max_col=ws.max_column): # Itera sobre todas as colunas
    max_lenght = 0 # Inicializando uma variável que pega o maior tamanho de cabeçalho
    coll_letter = col[0].column_letter # Pega a letra de cada coluna. Ex: A1 -> col[0] = A; B1 -> col[0] = B
    # A coleta da column_letter é necessária pois a função 'column_dimensions' não trabalha com numeros, so com letras

    for cell in col: # Itera sobre cada célula das colunas
        if cell.value: # Se a célula não for 'None'
            max_lenght = max(max_lenght, len(str(cell.value))) # Atualiza max_length com o maior tamanho de string encontrado na coluna

    ws.column_dimensions[coll_letter].width = max_lenght + 2 # Redimensiona a coluna para o tamanho da maior célula e adiciona 2 com espaçamento extra

wb.save('dados.xlsx')

"""
Preto	000000
Branco	FFFFFF
Vermelho	FF0000
Verde	008000
Azul	0000FF
Amarelo	FFFF00
Laranja	FFA500
Roxo	800080
Cinza	808080
"""

"""for col in ws.iter_cols(min_col=1, max_row=1):
    coll_letter = col[0].column_letter
    ws.column_dimensions[coll_letter].width = 16
    
Para caso eu queira contar a quantidade de caracteres da maior palavra do cabeçalho e colocar esse numero automaticamente nas dimensões
    """