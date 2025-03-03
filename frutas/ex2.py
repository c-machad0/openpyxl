from openpyxl import load_workbook

workbook = load_workbook('dados.xlsx')
worksheet = workbook.active

def atualiza_preco_produto(preco_unitario):
    return preco_unitario + 1.50

# Atualizando o preço de cada produto
for col in worksheet.iter_cols(min_col=3, min_row=2): # min_col = 3 corresponde à terceira coluna e min_row = 2 corresponde aos dados a partir da segunda linha
    for cell in col:
        cell.value = atualiza_preco_produto(cell.value)
        print(cell.value)

workbook.save('dados.xlsx')