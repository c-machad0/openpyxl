from openpyxl import load_workbook
from openpyxl.styles import numbers

wb = load_workbook('dados.xlsx')

for sheet in wb.sheetnames:
    if sheet in ['Vendas', 'Relatorio']:
        continue
    
    ws = wb[sheet]

    for row in range(2, 1 + ws.max_row):
        cell_preco = ws.cell(row=row, column=3)
        cell_preco.number_format = 'R$ #,##0.00' # Formato da moeda brasileira

        cell_total = ws.cell(row=row, column=4)
        cell_preco.number_format = 'R$ #,##0.00'

wb.save('dados.xlsx')
