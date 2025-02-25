from openpyxl import load_workbook

wb = load_workbook('dados.xlsx') # wb é a variável que contém o arquivo
ws = wb.active # ws é a variável que ativa a aba da planilha

col_produto = 1 # Coluna A
col_quantidade = 2 # Coluna B
col_preco_unitario = 3 # Coluna C

col_total_vendas = col_preco_unitario + 1

ws.cell(row=1, column=col_total_vendas, value='Total') # Adicionando nova coluna na planilha

max_row = ws.max_row # Armazena o numero máximo de linhas para que posteriormente possa ser iterado

for row in range(2, max_row + 1): # Itera a partir da segunda linha, pois a primeira é de cabeçalhos, até a ultima linha. 'max_range + 1' é porque o range sempre exclui o ultimo numero
    cell_quantidade = ws.cell(row=row, column=col_quantidade) # Armazena os dados da celula da coluna quantidade na linha atual da iteração
    cell_preco_unitario = ws.cell(row=row, column=col_preco_unitario) # Armazena os dados da celula da coluna preco_unitario na linha atual da iteração
    cell_total_vendas = ws.cell(row=row, column=col_total_vendas) # Armazena os dados da celula da coluna total_vendas na linha atual da iteração

    formula_total_vendas = f'={cell_quantidade.coordinate}*{cell_preco_unitario.coordinate}' # Pega o valor da celula atual de quantidade e multiplica pelo valor da celula atual do preço unitario
    # o coordinate serve para pegar o endereço dessa celula. Ex: B2 * C2
    cell_total_vendas.value = formula_total_vendas # Armazena o valor do cálcula entre as células

wb.save('dados.xlsx')