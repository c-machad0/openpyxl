from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

wb = load_workbook('dados.xlsx')

for sheet in wb.sheetnames:
    if sheet == 'Vendas' or sheet == 'Relatorio': # Evitando que acesse as planilhas de vendas ou Relatorio
        continue

    ws = wb[f'{sheet}'] # Acessando a aba atual da iteração

    for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=ws.max_column): # Iterando sobre todas as colunas das celulas
        for cell in col: # Iterando sobre todas as celular de cada coluna
            max_lenght = 0 # Inicializando variavel que vai pegar o tamanho da maior palavra
            col_letter = col[0].column_letter # Armazenando a letra da coluna atual do for. Ex: A1 -> col[0] = 'A'

            cell.font = Font(size=12, bold=True) # Altera a fonte para tamanho 12 e negrito
            cell.fill = PatternFill(start_color='FF0000', fill_type='solid') # Altera o preenchimento para cor vermelha
            cell.alignment = Alignment(horizontal='center', vertical='center') # Altera o alinhamento para o centro da célula

            if cell.value: # Verifica se a célula não é 'None'
                max_lenght = max(max_lenght, len(str(cell.value))) # Compara o valor atual de max_length com o tamanho da string atual e atualiza o max_lenght
                # Ex: max_lenght = 0 e a palavra é 'Produto'. len(produto) = 7. Ou seja, o maior valor comparando 0 e 7 é 7
    
        ws.column_dimensions[col_letter].width = max_lenght + 4 # Depois de analisar a coluna atual, aumenta o tamanho dela baseado no tamanho do titulo da coluna + 4 de espaçamento extra

wb.save('dados.xlsx')