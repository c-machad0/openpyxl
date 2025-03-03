from openpyxl import load_workbook

wb = load_workbook('dados.xlsx')

col_produto = 1 # Coluna A
col_quantidade = 2 # Coluna B
col_preco_unitario = 3 # Coluna C
col_total = 4 # Coluna D

for sheet in wb.sheetnames: # Iterando sobre as abas da planilha 'dados'
    if sheet == 'Vendas' and sheet == 'Relatorio': # Evitando que acesse as planilhas de vendas e relatorio
        continue
    ws = wb[sheet] # Ativando a planilha da iteração atual do laço

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4): # Iterando sobre todas as linhas até a ultima coluna
        linha = row[0].row # Como o iter_rows me dá uma tupla com todas as colunas da linha, utilizei o row[0] para pegar a celula da coluna quantidade
        # e o row[0].row para pegar o numero exato da celula

        cell_valor_quantidade = ws.cell(row=linha, column=col_quantidade) # Armazenando a célula de quantidade de cada linha
        cell_valor_preco = ws.cell(row=linha, column=col_preco_unitario) # Armazenando a célula de preço de cada linha
        cell_total = ws.cell(row=linha, column=col_total) # Armazenando a célula de total de cada linha

        formula_total = f'={cell_valor_quantidade.coordinate}*{cell_valor_preco.coordinate}' # =B1*C1
        cell_total.value = formula_total # Pegando o valor total do calculo de cada linha

wb.save('dados.xlsx')

       