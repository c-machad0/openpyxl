from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

wb = load_workbook('dados.xlsx')

# Se 'Relatorio' já existir, apenas pega a referência para ele
if 'Relatorio' in wb.sheetnames:
    ws = wb['Relatorio']
else:
    ws = wb.create_sheet('Relatorio')

headers = ['Total de vendas', 'Produto mais vendido', 'Produto mais rentável']

for id, header in enumerate(headers, start=1):
    ws.cell(row=1, column=id, value=header)

for col in ws.iter_cols(max_row=1, max_col=ws.max_column, min_col=1):
    coll_letter = col[0].column_letter
    ws.column_dimensions[coll_letter].width = 23

    for cell in col:
        cell.font = Font(size=12, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='FF0000', fill_type='solid')
        
wb.save('dados.xlsx')