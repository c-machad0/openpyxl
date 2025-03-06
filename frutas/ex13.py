from openpyxl import load_workbook

wb = load_workbook('dados.xlsx')

nome_produtos = [] # Inicializa uma lista que armazenará o nome dos produtos
ws = wb['Janeiro'] # Ativa a aba 'Janeiro' para pegar os produtos dela. Poderia ser de qualquer outra aba

for row in ws.iter_rows(min_col=1, max_col=1, min_row=2, max_row=ws.max_row): # Itera sobre as linhas
    for cell in row: # Itera sobre as celulas de cada linha
        nome_produtos.append(cell.value) # Adiciona o valor da célula, ou seja, nome do produto, na lista 'nome_produtos'

nome_produtos.sort() # Ordena os itens em ordem alfabética

# Pega a aba "Resumo"
sheet_resumo = wb['Resumo']

# Escreve cada produto em uma linha separada
# A ideia é escrever cada produto em uma linha diferente, usando enumerate() para associar corretamente cada produto a um índice de linha.
for idx, produto in enumerate(nome_produtos, start=2):  # Começa na linha 2
    sheet_resumo.cell(row=idx, column=1, value=produto)  # Escreve na coluna 1

wb.save('dados.xlsx')
        