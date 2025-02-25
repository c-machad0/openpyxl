from openpyxl import Workbook
import random

wb = Workbook()

worksheet = wb.active
worksheet.title = 'Vendas'

headers = ['Produto', 'Quantidade', 'Preço Unitário']

# Colocando cabeçalhos na planilha
for num_col, header in enumerate(headers, start=1): # Usei o enumerate para dar índices aos meus cabeçalhos, iniciando do 1, pois excel não existe indice 0
    worksheet.cell(row=1, column=num_col, value=header)

# Função que retorna o nome aleatório de produtos
def gerar_nome_produtos(produtos_disponiveis):
    return random.choice(produtos_disponiveis)

produtos = ['Uva', 'Morango', 'Maçã', 'Manga', 'Kiwi']

# Adicionando os itens por célula
for row in range(2, 2 + len(produtos)): # Inicio em 2, pois os itens começam na segunda linha e '2 + len(produtos)' significa a quantidade total de produtos + 2 linhas que ja foram adicionadas
    if not produtos: # Se não houver mais produtos na lista, o laço é quebrado
        break

    nome_produto = gerar_nome_produtos(produtos) # Busca um nome aleatório de produtos e salva na variavel
    produtos.remove(nome_produto) # Remove esse produto da lista, para não ser usado novamente

    qtd_produto = random.randint(1, 20)
    preco_produto = round(random.uniform(1, 10), 2)

    worksheet.cell(row=row, column=1, value=nome_produto)
    worksheet.cell(row=row, column=2, value=qtd_produto)
    worksheet.cell(row=row, column=3, value=preco_produto)

file_path = 'dados.xlsx'
wb.save(file_path)