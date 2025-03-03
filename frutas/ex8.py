from openpyxl import load_workbook
from openpyxl.chart import Reference, BarChart

wb = load_workbook('dados.xlsx') # Carrega a planilha
ws = wb['Vendas'] # Acessa os dados da aba vendas

values = Reference(ws, min_col=2, max_col=2, min_row=1, max_row=ws.max_row) # Captura os intervalos para os valores
cats = Reference(ws, min_col=1, max_col=1, min_row=2, max_row=ws.max_row) # Captura os intervalos para as categorias

chart = BarChart() # Instancia uma variavel na classe BarChart
chart.add_data(values, titles_from_data=True) # Adiciona os dados dos valores no gráfico e informa que o título não será lido como dado
chart.set_categories(cats) # Adiciona os dados das categorias no gráfico

chart.title = 'Quantidade de vendas por produto' # Titulo do gráfico
chart.x_axis.title = 'Produto' # Titulo do eixo x, onde estarão as categorias
chart.y_axis.title = 'Vendas por produto' # Titulo do eixo y, onde estarão a quantidade dos produtos

if 'Relatorio' in wb.sheetnames: # Verifica se a aba relatorio existe
    ws_destino = wb['Relatorio'] # Ativa a aba para ser usada

ws_destino.add_chart(chart, 'A8') # Adiciona o gráfico na aba 'Relatorio' na célula A8

wb.save('dados.xlsx')