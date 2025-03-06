from openpyxl import load_workbook

wb = load_workbook('dados.xlsx')

abas = ['Janeiro', 'Fevereiro', 'Março']
sheet_resumo = wb['Resumo']

# Criar um dicionário para armazenar a soma das quantidades por produto
quantidade_por_produto = {}

# Percorrer as abas e somar as quantidades por produto
for aba in abas:
    ws = wb[aba]
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
        produto = row[0].value  # Nome do produto (Coluna A)
        quantidade = row[1].value  # Quantidade (Coluna B)

        if isinstance(quantidade, (int, float)):  # Garantir que seja um número
            if produto in quantidade_por_produto: # Verifica se o produto ja existe no dicionário
                quantidade_por_produto[produto] += quantidade # Se existir, soma a quantidade atual do produto à nova quantidade
            else:
                quantidade_por_produto[produto] = quantidade # Caso não exista, o produto será adicionado como nova chave e a quantidade, será a atual da iteração

# Ordenar os produtos por nome
produtos_ordenados = sorted(quantidade_por_produto.keys())

# Escrever os produtos e seus totais na aba "Resumo"
linha_resumo = 2  # Começar a partir da linha 2
for produto in produtos_ordenados:
    sheet_resumo.cell(row=linha_resumo, column=2, value=quantidade_por_produto[produto])  # Total da quantidade
    linha_resumo += 1  # Passa para a próxima linha

wb.save('dados.xlsx')