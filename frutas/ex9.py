from openpyxl import load_workbook
from openpyxl.styles import PatternFill

wb = load_workbook('dados.xlsx')
ws = wb['Vendas']

for col in ws.iter_cols(min_col=4, max_col=4, min_row=2, max_row=ws.max_row): # Itera sobre a coluna 4 (Total)
    for cell in col: # Itera sobre as celulas da coluna 4
        if isinstance(cell.value, (int, float)) and cell.value < 100: # Verifica se o valor da celula é um numero int ou float e se esse numero é menor que 100
            cell.fill = PatternFill(start_color='FF0000', fill_type='solid') # Pinta de vermelho, todas as células que passem na verificação

wb.save('dados.xlsx')