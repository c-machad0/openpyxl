from openpyxl import load_workbook

wb = load_workbook('dados.xlsx', data_only=True)
ws = wb.active

for dados in ws.iter_cols(min_col=4, max_col=4, min_row=2, max_row=ws.max_row):
    for cell in dados:
        if cell.value and float(cell.value) > 50:
            print(f'Valores maiores que 50: {cell.value}')